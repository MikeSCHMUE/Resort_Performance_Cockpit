import numpy_financial as npf
import streamlit as st
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from fpdf import FPDF
from io import BytesIO
from openpyxl import Workbook
import seaborn as sns

st.set_page_config(page_title="ROI Dashboard", layout="centered")
st.markdown("""
<div style='text-align: center; margin-bottom: 20px;'>
    <h1 style='font-size: 2.8em;'>📊 Resort Performance Cockpit</h1>
</div>
""", unsafe_allow_html=True)

if "results" not in st.session_state:
    st.session_state["results"] = {}

with st.form("input_form"):
    st.subheader("🏘️ Accommodation Settings")
    rooms = st.number_input("Number of Rooms", min_value=1, value=5)
    days = st.number_input("Days per Month", min_value=1, value=30)

    st.subheader("⚙️ Season & Room Configuration")

    st.subheader("🌦️ High & Low Season Setup")
    high_months = st.slider("High Season Months", 0, 12, 4)
    low_months = 12 - high_months

    # Villa type configuration with sliders
    st.markdown("### 🏝️ Beachfront Villas")
    beach_rooms = st.number_input("Number of Beachfront Rooms", min_value=0, value=3)  # Added number input
    beach_high_price = st.slider("High Season Average Daily Rate (Beachfront)", 110, 180, 150, step=10)
    beach_low_price = st.slider("Low Season Average Daily Rate (Beachfront)", 90, 160, 120, step=10)
    beach_high_occ = st.slider("High Season Occupancy (Beachfront) %", 0, 100, 70, step=10) / 100
    beach_low_occ = st.slider("Low Season Occupancy (Beachfront) %", 0, 100, 50, step=10) / 100

    st.markdown("### 🏡 Garden View Villas")
    compact_rooms = st.number_input("Number of Garden View Rooms", min_value=0, value=2)  # Added number input
    compact_high_price = st.slider("High Season Average Daily Rate (Compact)", 80, 150, 120, step=10)
    compact_low_price = st.slider("Low Season Average Daily Rate (Compact)", 60, 130, 90, step=10)
    compact_high_occ = st.slider("High Season Occupancy (Compact) %", 0, 100, 70, step=10) / 100
    compact_low_occ = st.slider("Low Season Occupancy (Compact) %", 0, 100, 50, step=10) / 100

    st.subheader("💰 Revenue and Cost")
    ancillary = st.number_input("Ancillary Revenue per Night in USD (Scooter Rental, Laundry Service etc.)", value=10.0, step=5.0)
    fixed_costs = st.number_input("Monthly Fixed Costs in USD (OPEX = Operational Expenditures)", value=5500.0, step=500.0)
    variable_cost = st.number_input("Variable Cost per Room per Night in USD (Toiletries, Breakfast incl.)", value=15.0, step=5.0)
    capex = st.number_input("Total Investment in USD (CAPEX = Capital Expenditure)", value=350000.0, step=10000.0)
    
    years = st.slider("Forecast Years", 1, 12, 10)
    growth = st.slider("Annual Revenue Growth (%)", 0, 10, 3) / 100
    inflation = st.slider("Annual Cost Inflation (%)", 0, 10, 3) / 100
    submitted = st.form_submit_button("🚀 Calculate")

if submitted:
    # --- 1. Calculate Nights and Revenue ---
    # Error handling: Ensure non-negative values for rooms and days
    if rooms < 0:
        st.error("Number of rooms must be non-negative.")
        st.stop()  # Stop execution if input is invalid
    if days < 0:
        st.error("Days per month must be non-negative.")
        st.stop()

    # Beachfront Villas
    nights_high_beach = beach_rooms * days * high_months * beach_high_occ
    nights_low_beach = beach_rooms * days * low_months * beach_low_occ
    rev_beach = nights_high_beach * beach_high_price + nights_low_beach * beach_low_price

    # Compact Villas
    nights_high_compact = compact_rooms * days * high_months * compact_high_occ
    nights_low_compact = compact_rooms * days * low_months * compact_low_occ
    rev_compact = nights_high_compact * compact_high_price + nights_low_compact * compact_low_price

    # Gesamt
    nights_high = nights_high_beach + nights_high_compact
    nights_low = nights_low_beach + nights_low_compact
    total_nights = nights_high + nights_low

    rev_room = rev_beach + rev_compact
    rev_ancillary = total_nights * ancillary
    total_revenue = rev_room + rev_ancillary

    # --- 2. Calculate Costs and Profit ---
    total_opex = fixed_costs * 12 + total_nights * variable_cost
    profit = total_revenue - total_opex
    roi = (profit / capex * 100) if capex else 0  # Handle case where capex is zero
    ebitda = total_revenue - total_opex
    ebitda_margin = (ebitda / total_revenue) * 100 if total_revenue else 0

    # --- 3. Calculate Average Prices and Occupancy ---
    avg_price_beach = (beach_high_price * high_months + beach_low_price * low_months) / 12 if high_months + low_months else 0
    avg_price_compact = (compact_high_price * high_months + compact_low_price * low_months) / 12 if high_months + low_months else 0
    avg_price = (avg_price_beach * beach_rooms + avg_price_compact * compact_rooms) / (beach_rooms + compact_rooms) if (beach_rooms + compact_rooms) else 0
    avg_occ_beach = (beach_high_occ * high_months + beach_low_occ * low_months) / 12 if high_months + low_months else 0
    avg_occ_compact = (compact_high_occ * high_months + beach_low_occ * low_months) / 12 if high_months + low_months else 0
    avg_occ = (avg_occ_beach * beach_rooms + avg_occ_compact * compact_rooms) / (beach_rooms + compact_rooms) if (beach_rooms + compact_rooms) else 0

    # --- 4. Calculate Break-Even Point ---
    margin = (avg_price + ancillary) - variable_cost
    bep_nights = fixed_costs / margin if margin > 0 else None

    # --- 5. Multi-Year Forecast ---
    forecast = []
    cum_profit = 0
    payback = None
    for y in range(1, years + 1):
        rev = total_revenue * ((1 + growth) ** (y - 1))
        cost = total_opex * ((1 + inflation) ** (y - 1))
        prof = rev - cost
        cum_profit += prof
        roi_y = prof / capex * 100 if capex else 0
        if not payback and cum_profit >= capex:
            payback = y
        forecast.append([y, rev, prof, roi_y])
    df = pd.DataFrame(forecast, columns=["Year", "Revenue", "Profit", "ROI (%)"])

    # --- 6. Store Results in Session State ---
    st.session_state["results"] = {
        "nights_high": nights_high,
        "nights_low": nights_low,
        "total_nights": total_nights,
        "revenue_rooms": rev_room,
        "revenue_ancillary": rev_ancillary,
        "total_revenue": total_revenue,
        "total_costs": total_opex,
        "profit": profit,
        "roi": roi,
        "payback": payback,
        "df": df,
        "avg_price": avg_price,
        "avg_occ": avg_occ,
        "bep_nights": bep_nights,
        "capex": capex,
    "ebitda": ebitda,
    "ebitda_margin": ebitda_margin  # Store EBITDA info
    }

# --- 7. Display Results ---
if st.session_state["results"]:
    r = st.session_state["results"]
    bep_nights = r.get("bep_nights")

    st.markdown("<h2 style='text-align: center;'>📈 Annual Summary</h2>", unsafe_allow_html=True)

    # CSS Style für alle Boxen
    box_style = """
        border: 1px solid #ccc;
        padding: 10px;
        border-radius: 10px;
        text-align: center;
        box-shadow: 2px 2px 5px rgba(0,0,0,0.1);
        background-color: #f9f9f9;
        color: #111111;
        min-height: 180px;
        display: flex;
        flex-direction: column;
        justify-content: center;
    """

    # Obere 2 Spalten
    col1, col2 = st.columns(2)

    with col1:
        st.markdown(f"""
        <div style="{box_style}">
            <div><strong>High Season Nights Sold:</strong><br> {r['nights_high']:.0f}</div>
            <div><strong>Low Season Nights Sold:</strong><br> {r['nights_low']:.0f}</div>
            <div><strong>Total Nights Sold:</strong><br> {r['total_nights']:.0f}</div>
        </div>
        """, unsafe_allow_html=True)

    with col2:
        st.markdown(f"""
        <div style="{box_style}">
            <div><strong>Room Revenue:</strong><br> ${r['revenue_rooms']:,.2f}</div>
            <div><strong>Ancillary Revenue:</strong><br> ${r['revenue_ancillary']:,.2f}</div>
        </div>
        """, unsafe_allow_html=True)

    # Untere zentrierte Box
    st.markdown(f"""
    <div style='display: flex; justify-content: center; margin-top: 20px;'>
        <div style="{box_style}">
            <div style='font-size: 1.2em; font-weight: bold;'>Total Revenue:</div>
            <div style='font-size: 1.2em;'>${r['total_revenue']:,.2f}</div>
            <div style='margin-top: 10px; font-size: 1.2em; font-weight: bold;'>Total Costs:</div>
            <div style='font-size: 1.2em;'>${r['total_costs']:,.2f}</div>
            <div style='margin-top: 10px; font-size: 1.2em; font-weight: bold;'>Profit:</div>
            <div style='font-size: 1.2em;'>${r['profit']:,.2f}</div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    st.markdown("<div style='margin-top: 60px;'></div>", unsafe_allow_html=True)

    # --- 8. Heatmap ---
    st.subheader("🔥 ROI Heatmap")
    price_range = np.arange(80, 201, 20)
    occ_range = np.linspace(0.4, 1.0, 7)
    heatmap = []
    for p in price_range:
        row = []
        for o in occ_range:
            nh = rooms * days * high_months * o
            nl = rooms * days * low_months * o
            tn = nh + nl
            rev = nh * p + nl * p + tn * ancillary
            cost = tn * variable_cost + fixed_costs * 12
            roi_h = (rev - cost) / capex * 100 if capex else 0
            row.append(roi_h)
        heatmap.append(row)

    fig, ax = plt.subplots(figsize=(45, 30))
    sns.heatmap(
        heatmap,
        xticklabels=[f"{int(round(o*100))}%" for o in occ_range],
        yticklabels=[f"${p}" for p in price_range],
        cmap="YlGnBu",
        annot=True,
        fmt=".1f",
        ax=ax,
        annot_kws={"size": 55},
        cbar=False  # 👈 Colorbar ausschalten
    )
    #cbar = ax.collections[0].colorbar
    #cbar.ax.tick_params(labelsize=55)  # Change font size of colorbar labels
    #cbar.set_label("ROI (%)", fontsize=55)  # Change font size of colorbar title
    try:
        x_idx = min(range(len(occ_range)), key=lambda i: abs(occ_range[i] - r["avg_occ"]))
        y_idx = min(range(len(price_range)), key=lambda i: abs(price_range[i] - r["avg_price"]))
        ax.add_patch(plt.Rectangle((x_idx, y_idx), 1, 1, fill=False, edgecolor='red', lw=8))
    except ValueError:
        pass
    ax.set_xlabel("Occupancy Rate", fontsize=60, labelpad=30)
    ax.set_ylabel("Room Price", fontsize=60, labelpad=30)
    ax.set_title("ROI by Price and Occupancy", fontsize=85, pad=65)
    ax.tick_params(axis='both', labelsize=55)  # <- das ist neu
    col_hm = st.columns([0.15, 0.7, 0.15])[1]
    with col_hm:
        st.pyplot(fig)
        
        # Beschriftung mittig zentriert
    st.markdown(
        "<div style='text-align: center; margin-top: -15px;'>"
        "<span style='color: gray;'>ℹ️ ROI = Estimated Year 1 Return on Investment – before inflation and revenue growth.</span>"
        "</div>",
        unsafe_allow_html=True
    )

    st.markdown("<div style='margin-top: 60px;'></div>", unsafe_allow_html=True)   

    # --- 9. Break-Even Chart ---

    st.subheader("📉 Break-Even Chart")
    nights_range = np.arange(0, rooms * days + 10, 5)
    rev_line = nights_range * r["avg_price"] + nights_range * ancillary
    cost_line = fixed_costs + nights_range * variable_cost
    fig2, ax2 = plt.subplots(figsize=(16, 12))
    ax2.plot(nights_range, rev_line, label="Revenue", linewidth=2, markersize=8)
    ax2.plot(nights_range, cost_line, label="Total Cost", linewidth=2, markersize=8)
    if bep_nights:
        ax2.axvline(bep_nights, color="red", linestyle="--", linewidth=2, label=f"Break-Even ({bep_nights:.0f} Nights)")

        max_nights = rooms * days
        break_even_occ = bep_nights / max_nights * 100

        ax2.text(
            bep_nights + 14,
            -300,  # unterhalb der X-Achse (je nach Skalierung evtl. -500 testen)
            f"{break_even_occ:.1f}% Occ.",
            color="red",
            fontsize=20,
            fontweight="bold",
            horizontalalignment="center"
        )

    else:
        ax2.text(0.5, 0.5, "Break-even point cannot be calculated\n(non-positive margin)", ha="center", va="center", fontsize=12, color="gray")
    ax2.set_title("Break-Even Analysis", fontsize=30, pad=25)
    ax2.set_xlabel("Nights Sold", fontsize=26,labelpad=10)
    ax2.set_ylabel("USD", fontsize=26,labelpad=10)
    # Manuelle Positionierung (optional)
    ax2.xaxis.set_label_coords(0.5, -0.1)  # 0.5 = Mitte; -0.1 = weiter nach unten
    ax2.yaxis.set_label_coords(-0.1, 0.5)  # -0.1 = weiter nach links
    ax2.legend(fontsize=18)
    ax2.tick_params(axis='both', labelsize=16)

    # Nach links verschoben, wie bei der Heatmap
    col_bep = st.columns([0.08, 0.8, 0.15])[1]
    with col_bep:
        st.pyplot(fig2)

    st.markdown("<div style='margin-top: 40px;'></div>", unsafe_allow_html=True)

    # --- 10. Multi-Year Forecast Table ---
    st.subheader("📆 Multi-Year Forecast")
    df_styled = r["df"].style.apply(
        lambda x: ["font-weight: bold" if x.name in ["Year", "Revenue", "Profit", "ROI (%)"] else "" for i in x], axis=1
    )
    st.dataframe(r["df"].style.format({"Revenue": "${:,.0f}", "Profit": "${:,.0f}", "ROI (%)": "{:.1f}"}))

    if r["payback"]:
        st.info(f"💸 Investment recovered in **Year {r['payback']}**.")
    else:
        st.warning("⚠️ Investment not recovered during forecast period.")
    
    st.markdown("<div style='margin-top: 50px;'></div>", unsafe_allow_html=True)

    # --- 13. Additional Investor Metrics ---
    st.subheader("📊 Investor Insights")

    # Cumulative Cash Flow Chart
    cumulative_profits = np.cumsum(r["df"]["Profit"])
    years = r["df"]["Year"]

    fig3, ax3 = plt.subplots(figsize=(12, 6))
    ax3.plot(years, cumulative_profits, marker="o", linestyle="-", linewidth=2, label="Cumulative Profit")
    ax3.axhline(y=r["capex"], color='red', linestyle='--', linewidth=2, label="CAPEX (Investment)")
    ax3.set_title("Cumulative Cash Flow", fontsize=20)
    ax3.set_xlabel("Year", fontsize=16)
    ax3.set_ylabel("USD", fontsize=16)
    ax3.legend(fontsize=12)
    ax3.grid(True)
    st.pyplot(fig3)

    st.markdown("<div style='margin-top: 40px;'></div>", unsafe_allow_html=True)

# --- 14. Key Investment Metrics Table ---

if "results" in st.session_state:
    r = st.session_state["results"]

    capex_value = r.get("capex", 0)
    df_profit = r.get("df")

    if df_profit is not None and "Profit" in df_profit.columns:
        cash_flows = [-capex_value] + df_profit["Profit"].tolist()

        irr = None
        try:
            irr = npf.irr(cash_flows)
        except Exception:
            pass

        # Neu: NPV berechnen
        discount_rate = 0.08
        npv = None
        try:
            npv = npf.npv(discount_rate, cash_flows)  # Changed this line
        except Exception:
            pass

        st.subheader("📊 Key Investment Metrics")

        # Bewertung für Score-Spalte
        def score_irr(val): return "🟢" if val > 0.10 else "🟠" if val > 0.05 else "🔴"
        def score_npv(val): return "🟢" if val > 0 else "🟠" if val == 0 else "🔴"
        def score_payback(val): return "🟢" if val <= 7 else "🟠" if val <= 11 else "🔴"
        def score_bep(nights):
            if nights is None:
                return "N/A"
            elif nights <= 50:
                return "🟢"
            elif nights <= 80:
                return "🟠"
            else:
                return "🔴"
        def score_roi(roi_value):
            if roi_value is None:
                return "N/A"
            elif roi_value > 18:
                return "🟢"
            elif 12 <= roi_value <= 18:
                return "🟠"
            else:
                return "🔴"

        # Score-Werte berechnen
        score_irr_val = score_irr(irr) if irr is not None else "N/A"
        score_npv_val = score_npv(npv) if npv is not None else "N/A"
        score_payback_val = score_payback(r["payback"]) if r.get("payback") else "N/A"
        score_bep_val = score_bep(r["bep_nights"]) if r.get("bep_nights") is not None else "N/A"
        score_roi_val = score_roi(r["roi"]) if r.get("roi") is not None else "N/A" # Calculate ROI score

        
        metrics_data = {
            "Metric": [
                "Return on Investment (ROI)",
                "Break-Even Point (Nights)",
                "Internal Rate of Return (IRR)",
                "Net Present Value (NPV)",
                "Payback Period",
                "EBITDA Margin"
            ],
            "Value": [
                f"{r['roi']:.1f}%" if r.get("roi") is not None else "N/A",
                f"{r['bep_nights']:.0f} Nights" if r.get("bep_nights") is not None else "N/A",
                f"{irr * 100:.1f}%" if irr is not None else "N/A",
                f"${npv:,.0f}" if npv is not None else "N/A",
                f"{r['payback']} Years" if r.get("payback") else "Not reached",
                f"{r['ebitda_margin']:.1f}%" if r.get("ebitda_margin") is not None else "N/A"
            ],
            "Score": [
                score_roi_val,
                score_bep_val,
                score_irr_val,
                score_npv_val,
                score_payback_val,
                "🟢" if r.get("ebitda_margin", 0) >= 30 else "🟠" if r.get("ebitda_margin", 0) >= 15 else "🔴"
            ]
        }

        df_metrics = pd.DataFrame(metrics_data)
        df_metrics.index = df_metrics.index + 1
        st.table(df_metrics)
    else:
        st.warning("⚠️ Forecast data (Profit) is missing. Run calculation first.")
else:
    st.info("ℹ️ Please submit your inputs to see financial metrics.")

# --- 15. Final Combined Export (Excel + PDF Enhanced) ---

if "results" in st.session_state:
    r = st.session_state["results"]
    if all(k in r for k in ["nights_high", "nights_low", "total_nights", "revenue_rooms", "revenue_ancillary", "total_revenue", "total_costs", "profit", "roi", "payback", "bep_nights", "df", "capex"]):
        irr = None
        npv = None
        try:
            irr = npf.irr([-r["capex"]] + r["df"]["Profit"].tolist())
            npv = npf.npv(0.08, [-r["capex"]] + r["df"]["Profit"].tolist())
        except Exception:
            pass

        st.markdown("<div style='margin-top: 40px;'></div>", unsafe_allow_html=True)

        # Excel Export
        final_buffer = BytesIO()
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Annual Summary"

        summary_data = [
            ("High Season Nights", f"{r['nights_high']:.0f}"),
            ("Low Season Nights", f"{r['nights_low']:.0f}"),
            ("Total Nights", f"{r['total_nights']:.0f}"),
            ("Room Revenue", f"${r['revenue_rooms']:,.0f}"),
            ("Ancillary Revenue", f"${r['revenue_rooms']:,.0f}"),
            ("Total Revenue", f"${r['revenue_rooms']:,.0f}"),
            ("Total Costs", f"${r['revenue_rooms']:,.0f}"),
            ("Annual Profit", f"${r['revenue_rooms']:,.0f}"),
        ]

        for i, (label, value) in enumerate(summary_data, 1):
            ws1[f"A{i}"] = label
            ws1[f"B{i}"] = value

        ws2 = wb.create_sheet("Forecast")
        for j, col in enumerate(r["df"].columns, 1):
            ws2.cell(row=1, column=j, value=col)
        for i, row in enumerate(r["df"].itertuples(index=False), 2):
            for j, val in enumerate(row, 1):
                ws2.cell(row=i, column=j, value=val)

        ws3 = wb.create_sheet("Key Metrics")
        
        key_metrics = [
            ("Return on Investment (ROI)", f"{r['roi']:.1f}%"),
            ("Break-Even Point (Nights)", f"{r['bep_nights']:.0f}" if r["bep_nights"] else "N/A"),
            ("Internal Rate of Return (IRR)", f"{irr * 100:.1f}%" if irr else "N/A"),
            ("Net Present Value (NPV)", f"${npv:,.0f}" if npv else "N/A"),
            ("Payback Period", f"{r['payback']} Years" if r["payback"] else "Not reached"),
            ("EBITDA Margin", f"{r['ebitda_margin']:.1f}%" if r.get("ebitda_margin") else "N/A")
        ]


        for i, (label, value) in enumerate(key_metrics, 1):
            ws3[f"A{i}"] = label
            ws3[f"B{i}"] = value

        wb.save(final_buffer)
        final_buffer.seek(0)
        st.download_button("📥 Download Final Excel Report", data=final_buffer,
                        file_name="resort_dashboard_report.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        # PDF Export
        class PDF(FPDF):
            def header(self):
                try:
                    self.image("logo.png", x=160, y=8, w=40)
                except RuntimeError:
                    pass
                self.set_font("Arial", "B", 14)
                self.cell(0, 10, "Resort Investment Report", ln=True, align="C")
                self.ln(10)

            def footer(self):
                self.set_y(-15)
                self.set_font("Arial", "I", 8)
                self.cell(0, 10, f"Page {self.page_no()}", align="R")

            def title_page(self, project_name="Beachfront Eco Resort"):
                self.add_page()
                self.set_font("Arial", "B", 24)
                self.cell(0, 80, "", ln=True)
                self.cell(0, 10, project_name, ln=True, align="C")
                self.set_font("Arial", "", 16)
                self.cell(0, 10, "Investment Performance Report", ln=True, align="C")
                self.ln(20)

            def section(self, title, content_lines):
                self.set_font("Arial", "B", 12)
                self.cell(0, 10, title, ln=True)
                self.set_font("Arial", "", 11)
                for line in content_lines:
                    self.cell(0, 8, str(line), ln=True)
                self.ln(5)

        if st.button("📄 Generate PDF Report"):
            pdf = PDF()
            pdf.set_auto_page_break(auto=True, margin=15)
            pdf.title_page(project_name="Beachfront Resort - Puru Kambera")
            pdf.add_page()

            # 📌 Eingabewerte für PDF definieren
            inputs_lines = [
            f"Beachfront Villas: {beach_rooms} rooms",
            f"Garden View Villas: {compact_rooms} rooms",
            f"High Season Months: {high_months} / Low Season Months: {low_months}",
            f"Beachfront ADR (High/Low): ${beach_high_price} / ${beach_low_price}",
            f"Compact ADR (High/Low): ${compact_high_price} / ${compact_low_price}",
            f"Occupancy Beach (High/Low): {int(beach_high_occ*100)}% / {int(beach_low_occ*100)}%",
            f"Occupancy Compact (High/Low): {int(compact_high_occ*100)}% / {int(compact_low_occ*100)}%",
            f"Ancillary Revenue per Night: ${ancillary:.2f}",
            f"Monthly OPEX: ${fixed_costs:.2f}",
            f"Variable Cost per Room/Night: ${variable_cost:.2f}",
            f"Initial Investment (CAPEX): ${capex:,.0f}",
            f"Annual Revenue Growth: {int(growth * 100)}%",
            f"Annual Cost Inflation: {int(inflation * 100)}%"
            ]
            pdf.section("Input", inputs_lines)

            pdf.add_page()

            summary_lines = [f"{label}: {value}" for label, value in summary_data]
            forecast_lines = [f"Year {int(row[0])}: Revenue ${row[1]:,.0f}, Profit ${row[2]:,.0f}, ROI {row[3]:.1f}%"
                            for row in r["df"].itertuples(index=False)]
            key_metric_lines = [f"{label}: {value}" for label, value in key_metrics]
            pdf.section("Annual Summary", summary_lines)
            pdf.section("Forecast", forecast_lines)
            pdf.section("Key Metrics", key_metric_lines)

            pdf_output = pdf.output(dest="S").encode("latin-1")
            st.download_button("📄 Download Final PDF Report", data=pdf_output,
                            file_name="resort_dashboard_report.pdf",
                            mime="application/pdf")
